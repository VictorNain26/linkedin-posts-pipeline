"""Tests pour import_analytics_csv : parser XLSX LinkedIn FR 2026.

Fixtures : on construit des XLSX synthétiques via openpyxl (zéro dépendance externe).
"""

import openpyxl
import pytest


def _build_test_xlsx(path):
    """Crée un XLSX synthétique avec la structure LinkedIn FR (5 feuilles)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # vire la feuille par défaut

    # DÉCOUVERTE
    sh = wb.create_sheet("DÉCOUVERTE")
    sh["A1"] = "Performance globale"
    sh["B1"] = "18/05/2026 - 24/05/2026"
    sh["A2"] = "Impressions"
    sh["B2"] = 178

    # ENGAGEMENT
    sh = wb.create_sheet("ENGAGEMENT")
    sh["A1"] = "Date"
    sh["B1"] = "Impressions"
    sh["C1"] = "Interactions"
    sh["A2"] = "18/05/2026"
    sh["B2"] = 64
    sh["C2"] = 0

    # MEILLEURS POSTS — 2 tableaux côte à côte, header sur row 3
    sh = wb.create_sheet("MEILLEURS POSTS")
    sh["A1"] = "50 posts disponibles au maximum"
    sh["A3"] = "URL du post"
    sh["B3"] = "Date de publication du post"
    sh["C3"] = "Interactions"
    sh["E3"] = "URL du post"
    sh["F3"] = "Date de publication du post"
    sh["G3"] = "Impressions"
    # Row 4 : un post avec interactions ET impressions (col 4=E)
    sh["A4"] = "https://www.linkedin.com/feed/update/urn:li:activity:7234567890123456789/"
    sh["B4"] = "22/05/2026"
    sh["C4"] = 5
    sh["E4"] = "https://www.linkedin.com/feed/update/urn:li:activity:7234567890123456789/"
    sh["F4"] = "22/05/2026"
    sh["G4"] = 51
    # Row 5 : un autre post juste avec impressions
    sh["E5"] = "https://www.linkedin.com/feed/update/urn:li:activity:9876543210987654321/"
    sh["F5"] = "20/05/2026"
    sh["G5"] = 29

    # ABONNÉS
    sh = wb.create_sheet("ABONNÉS")
    sh["A1"] = "Nombre total d'abonnés le 24/05/2026"
    sh["B1"] = 1005
    sh["A3"] = "Date"
    sh["B3"] = "Nouveaux abonnés"
    sh["A4"] = "18/05/2026"
    sh["B4"] = 3
    sh["A5"] = "19/05/2026"
    sh["B5"] = 7

    # DONNÉES DÉMOGRAPHIQUES
    sh = wb.create_sheet("DONNÉES DÉMOGRAPHIQUES")
    sh["A1"] = "Principales données démographiques"
    sh["B1"] = "Valeur"
    sh["C1"] = "Pourcentage"
    sh["A2"] = "Lieux"
    sh["B2"] = "Paris et périphérie"
    sh["C2"] = 0.511
    sh["A3"] = "Lieux"
    sh["B3"] = "Lyon"
    sh["C3"] = 0.044

    wb.save(path)


class TestParseTopPostsSheet:
    def test_parser_extracts_posts(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _build_test_xlsx(xlsx)

        from import_analytics_csv import _parse_top_posts_sheet
        posts = _parse_top_posts_sheet(xlsx, "MEILLEURS POSTS")

        # 2 posts uniques (7234... avec interactions + impressions, 9876... avec impressions seules)
        activity_ids = {p["activity_id"] for p in posts}
        assert "7234567890123456789" in activity_ids
        assert "9876543210987654321" in activity_ids

        # Le post 7234 doit avoir les 2 métriques
        p1 = next(p for p in posts if p["activity_id"] == "7234567890123456789")
        assert p1["impressions"] == 51
        assert p1["interactions"] == 5
        assert p1["date"] == "2026-05-22"


class TestParseFollowersSheet:
    def test_parser_extracts_growth_and_total(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _build_test_xlsx(xlsx)

        from import_analytics_csv import _parse_followers_sheet
        rows, total = _parse_followers_sheet(xlsx, "ABONNÉS")

        assert total == 1005
        assert ("2026-05-18", 3) in rows
        assert ("2026-05-19", 7) in rows


class TestParseDemographicsSheet:
    def test_parser_extracts_demo_rows(self, tmp_path):
        xlsx = tmp_path / "test.xlsx"
        _build_test_xlsx(xlsx)

        from import_analytics_csv import _parse_demographics_sheet
        rows = _parse_demographics_sheet(xlsx, "DONNÉES DÉMOGRAPHIQUES")

        # rows = list de (dimension, value, percentage)
        lieux = [r for r in rows if r[0] == "Lieux"]
        assert len(lieux) == 2
        # Paris devrait être présent avec 51.1%
        paris = next(r for r in lieux if "Paris" in r[1])
        assert paris[2] == pytest.approx(0.511)


class TestExtractActivityId:
    def test_extracts_from_standard_url(self):
        from import_analytics_csv import _extract_activity_id
        url = "https://www.linkedin.com/feed/update/urn:li:activity:7234567890123456789/"
        assert _extract_activity_id(url) == "7234567890123456789"

    def test_extracts_from_ugc_url(self):
        from import_analytics_csv import _extract_activity_id
        url = "https://www.linkedin.com/posts/foo-activity-9999999999999999999-abc"
        assert _extract_activity_id(url) == "9999999999999999999"

    def test_returns_none_on_invalid(self):
        from import_analytics_csv import _extract_activity_id
        assert _extract_activity_id("") is None
        assert _extract_activity_id(None) is None
        assert _extract_activity_id("https://example.com/no-id-here") is None


class TestImportXlsxEndToEnd:
    def test_full_import(self, tmp_path, tmp_data_dir):
        xlsx = tmp_path / "test.xlsx"
        _build_test_xlsx(xlsx)

        from import_analytics_csv import import_xlsx
        summary = import_xlsx(xlsx)

        # Aucun match avec DB existante → tous créés comme external
        assert summary["posts_external_created"] >= 2
        assert summary["metrics_written"] >= 2
        assert summary["follower_days_imported"] == 2
        assert summary["demo_rows_imported"] == 2
